from flask import Flask, send_from_directory, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date
import time
import os
import logging
import webbrowser
import threading

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__, static_folder='.')
CORS(app)

EXCEL_FILE = "Work Tracker 2026.xlsx"
YEAR = 2026

# ── timer state ──────────────────────────────────────────────
# accumulated_seconds: time banked from before the current start (for pause/resume)
timer_state = {
    "start_time": None,
    "active_category": None,
    "paused": False,
    "accumulated_seconds": 0.0
}

# ── helpers ──────────────────────────────────────────────────

def day_col_for_date(ws, target: date):
    label = f"{target.strftime('%a')}\n{target.month}/{target.day}"
    for col in range(2, ws.max_column + 2):
        if ws.cell(row=1, column=col).value == label:
            return col
    return None

def get_categories(ws):
    cats = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            cats.append(str(val).strip())
    return cats

def find_category_row(ws, category):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == category:
            return row
    return None

def add_category_row(ws, category):
    new_row = ws.max_row + 1
    thin = Side(style="thin", color="e5e7eb")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    is_even = (new_row % 2 == 0)
    fill = PatternFill("solid", start_color="f8f9fa" if is_even else "ffffff")

    cat_cell = ws.cell(row=new_row, column=1)
    cat_cell.value = category
    cat_cell.font = Font(name="Calibri", bold=True, size=10, color="1a1a2e")
    cat_cell.fill = fill
    cat_cell.alignment = Alignment(horizontal="left", vertical="center")
    cat_cell.border = border
    ws.row_dimensions[new_row].height = 18

    for i in range(365):
        col = i + 2
        cell = ws.cell(row=new_row, column=col)
        cell.value = 0
        cell.number_format = '[h]:mm:ss'
        cell.font = Font(name="Calibri", size=9, color="374151")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = border

    logging.info(f"Added new category: '{category}' at row {new_row}")
    return new_row

def get_cell_seconds(cell):
    """Parse cell value into total seconds — handles all types openpyxl may return."""
    val = cell.value
    if val is None or val == "" or val == 0:
        return 0.0
    # openpyxl reads back time-formatted cells as timedelta objects
    if isinstance(val, timedelta):
        return val.total_seconds()
    if isinstance(val, (int, float)):
        return float(val) * 86400  # Excel day fraction
    if hasattr(val, 'hour'):  # time object
        return val.hour * 3600 + val.minute * 60 + val.second
    try:
        parts = str(val).strip().split(":")
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except Exception:
        pass
    return 0.0

def fmt(seconds):
    s = int(seconds)
    return f"{s//3600}:{(s%3600)//60:02}:{s%60:02}"

def current_elapsed():
    """Total elapsed seconds including any accumulated paused time."""
    acc = timer_state["accumulated_seconds"]
    if timer_state["start_time"] and not timer_state["paused"]:
        acc += time.time() - timer_state["start_time"]
    return acc

def update_excel(category, target_date, elapsed_seconds):
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"'{EXCEL_FILE}' not found.")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    col = day_col_for_date(ws, target_date)
    if col is None:
        raise ValueError(f"No column found for {target_date}")

    row = find_category_row(ws, category)
    is_new = row is None
    if is_new:
        row = add_category_row(ws, category)

    cell = ws.cell(row=row, column=col)
    existing = get_cell_seconds(cell)
    total = existing + elapsed_seconds

    cell.value = total / 86400
    cell.number_format = '[h]:mm:ss'
    cell.alignment = Alignment(horizontal="right", vertical="center")

    wb.save(EXCEL_FILE)
    logging.info(f"Updated '{category}' on {target_date}: +{fmt(elapsed_seconds)}, total={fmt(total)}")
    return fmt(total), is_new

# ── routes ────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/api/categories', methods=['GET'])
def api_categories():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        cats = get_categories(ws)
        return jsonify({"categories": cats})
    except Exception as e:
        return jsonify({"categories": [], "error": str(e)})

@app.route('/api/start', methods=['POST'])
def api_start():
    data = request.json
    category = data.get('category', '').strip()

    if not category:
        return jsonify({"success": False, "message": "Category cannot be empty."}), 400

    if timer_state["active_category"]:
        return jsonify({
            "success": False,
            "message": f"Timer already running for '{timer_state['active_category']}'. Stop it first."
        }), 400

    timer_state["start_time"] = time.time()
    timer_state["active_category"] = category
    timer_state["paused"] = False
    timer_state["accumulated_seconds"] = 0.0

    today = datetime.today().strftime("%A, %B %d").replace(" 0", " ")
    logging.info(f"Started timer for '{category}'")
    return jsonify({"success": True, "category": category, "day": today})

@app.route('/api/pause', methods=['POST'])
def api_pause():
    if not timer_state["active_category"]:
        return jsonify({"success": False, "message": "No timer running."}), 400
    if timer_state["paused"]:
        return jsonify({"success": False, "message": "Already paused."}), 400

    timer_state["accumulated_seconds"] += time.time() - timer_state["start_time"]
    timer_state["start_time"] = None
    timer_state["paused"] = True

    logging.info(f"Paused '{timer_state['active_category']}' at {fmt(timer_state['accumulated_seconds'])}")
    return jsonify({
        "success": True,
        "elapsed": fmt(timer_state["accumulated_seconds"]),
        "elapsed_seconds": int(timer_state["accumulated_seconds"])
    })

@app.route('/api/resume', methods=['POST'])
def api_resume():
    if not timer_state["active_category"]:
        return jsonify({"success": False, "message": "No timer running."}), 400
    if not timer_state["paused"]:
        return jsonify({"success": False, "message": "Timer is not paused."}), 400

    timer_state["start_time"] = time.time()
    timer_state["paused"] = False

    logging.info(f"Resumed '{timer_state['active_category']}'")
    return jsonify({"success": True, "category": timer_state["active_category"]})

@app.route('/api/stop', methods=['POST'])
def api_stop():
    if not timer_state["active_category"]:
        return jsonify({"success": False, "message": "No timer running."}), 400

    elapsed = current_elapsed()
    category = timer_state["active_category"]
    today = date.today()

    try:
        new_total, is_new = update_excel(category, today, elapsed)
        timer_state["start_time"] = None
        timer_state["active_category"] = None
        timer_state["paused"] = False
        timer_state["accumulated_seconds"] = 0.0
        return jsonify({
            "success": True,
            "category": category,
            "duration": fmt(elapsed),
            "new_total": new_total,
            "is_new_category": is_new
        })
    except Exception as e:
        logging.error(f"Stop error: {e}", exc_info=True)
        timer_state["start_time"] = None
        timer_state["active_category"] = None
        timer_state["paused"] = False
        timer_state["accumulated_seconds"] = 0.0
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/status', methods=['GET'])
def api_status():
    if timer_state["active_category"]:
        elapsed = current_elapsed()
        return jsonify({
            "active": True,
            "paused": timer_state["paused"],
            "category": timer_state["active_category"],
            "elapsed": fmt(elapsed),
            "elapsed_seconds": int(elapsed)
        })
    return jsonify({"active": False, "paused": False})

@app.route('/api/today', methods=['GET'])
def api_today():
    """Read today's totals directly from the spreadsheet."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        col = day_col_for_date(ws, date.today())
        if col is None:
            return jsonify({"totals": []})
        totals = []
        for row in range(2, ws.max_row + 1):
            cat = ws.cell(row=row, column=1).value
            if not cat:
                continue
            secs = get_cell_seconds(ws.cell(row=row, column=col))
            if secs > 0:
                totals.append({"category": str(cat), "total": fmt(secs), "seconds": int(secs)})
        totals.sort(key=lambda x: x["seconds"], reverse=True)
        return jsonify({
            "totals": totals,
            "date": date.today().strftime("%A, %B %d").replace(" 0", " ")
        })
    except Exception as e:
        logging.error(f"api_today error: {e}", exc_info=True)
        return jsonify({"totals": [], "error": str(e)})

def open_browser():
    time.sleep(1.5)
    webbrowser.open('http://localhost:5000')

if __name__ == '__main__':
    print(f"\n{'='*50}")
    print(f"  Work Tracker 2026")
    print(f"  http://localhost:5000")
    print(f"{'='*50}\n")
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(debug=True, port=5000, use_reloader=False)