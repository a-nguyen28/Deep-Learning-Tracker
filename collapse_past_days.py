"""
Run this once to collapse all past-day columns in your existing tracker.
Also safe to re-run anytime to update which days are collapsed.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date

EXCEL_FILE = "Work Tracker 2026.xlsx"

def collapse_past_days():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    today = date.today()

    today_label = f"{today.strftime('%a')}\n{today.month}/{today.day}"

    today_col = None
    last_col = ws.max_column

    for col in range(2, last_col + 1):
        if ws.cell(row=1, column=col).value == today_label:
            today_col = col
            break

    if today_col is None:
        print(f"Couldn't find today ({today}) in the header row. Is this a 2026 tracker?")
        return

    collapsed = 0
    for col in range(2, today_col):
        cl = get_column_letter(col)
        ws.column_dimensions[cl].outlineLevel = 1
        ws.column_dimensions[cl].hidden = True
        collapsed += 1

    # Make sure today onward is visible
    for col in range(today_col, last_col + 1):
        cl = get_column_letter(col)
        ws.column_dimensions[cl].outlineLevel = 0
        ws.column_dimensions[cl].hidden = False

    # [+] button appears on the LEFT of the group
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    wb.save(EXCEL_FILE)
    print(f"Done. Collapsed {collapsed} past-day columns. Today starts at column {get_column_letter(today_col)}.")

if __name__ == "__main__":
    collapse_past_days()
