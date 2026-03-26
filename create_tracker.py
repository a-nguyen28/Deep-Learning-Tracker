import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

YEAR = 2026
OUTPUT_FILE = "Work Tracker 2026.xlsx"

INITIAL_CATEGORIES = [
    "Classes",
    "Personal",
    "Coursework",
    "Employment",
    "UAV Lab",
    "Machine Learning",
    "Combat Clubs"
]

def make_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026"

    # Generate all days of year
    start = date(YEAR, 1, 1)
    days = [start + timedelta(days=i) for i in range(365 if YEAR != 2024 else 366)]

    # --- Header row ---
    # A1 = "Category", then one column per day
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    cat_fill = PatternFill("solid", start_color="1a1a2e")
    day_fill_weekday = PatternFill("solid", start_color="16213e")
    day_fill_weekend = PatternFill("solid", start_color="0f3460")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Category column (col 1)
    ws.column_dimensions["A"].width = 18
    ws["A1"].value = "Category"
    ws["A1"].font = header_font
    ws["A1"].fill = cat_fill
    ws["A1"].alignment = center

    for i, d in enumerate(days):
        col = i + 2  # starts at column B
        col_letter = get_column_letter(col)
        cell = ws.cell(row=1, column=col)
        # Format: "Mon\n1/1"
        day_label = f"{d.strftime('%a')}\n{d.month}/{d.day}"
        cell.value = day_label
        cell.font = header_font
        cell.alignment = center
        is_weekend = d.weekday() >= 5
        cell.fill = day_fill_weekend if is_weekend else day_fill_weekday
        ws.column_dimensions[col_letter].width = 5.5
        ws.row_dimensions[1].height = 30

    # --- Category rows ---
    row_fill_even = PatternFill("solid", start_color="f8f9fa")
    row_fill_odd = PatternFill("solid", start_color="ffffff")
    cat_label_font = Font(name="Calibri", bold=True, size=10, color="1a1a2e")
    time_font = Font(name="Calibri", size=9, color="374151")
    thin = Side(style="thin", color="e5e7eb")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, cat in enumerate(INITIAL_CATEGORIES):
        row = r_idx + 2
        # Category name cell
        cat_cell = ws.cell(row=row, column=1)
        cat_cell.value = cat
        cat_cell.font = cat_label_font
        cat_cell.fill = row_fill_even if r_idx % 2 == 0 else row_fill_odd
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        cat_cell.border = border
        ws.row_dimensions[row].height = 18

        for col in range(2, len(days) + 2):
            cell = ws.cell(row=row, column=col)
            cell.value = 0
            cell.number_format = '[h]:mm:ss'
            cell.font = time_font
            cell.fill = row_fill_even if r_idx % 2 == 0 else row_fill_odd
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border

    # Freeze pane at B2 so category col + header row stay visible
    ws.freeze_panes = "B2"

    wb.save(OUTPUT_FILE)
    print(f"Created {OUTPUT_FILE} with {len(days)} day columns and {len(INITIAL_CATEGORIES)} categories.")

if __name__ == "__main__":
    make_tracker()
