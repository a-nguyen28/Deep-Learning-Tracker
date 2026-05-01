"""
Work Tracker 2026 — Analysis Script
Run: python analyze_tracker.py
Expects "Work Tracker 2026.xlsx" in the same directory.
"""

import openpyxl
import pandas as pd
from datetime import timedelta, date
from collections import defaultdict

EXCEL_FILE = "Work Tracker 2026.xlsx"

# ── Load & parse ─────────────────────────────────────────────

def load_data(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Parse column headers → dates
    col_dates = {}
    for col in range(2, ws.max_column + 1):
        header = ws.cell(1, col).value
        if not header:
            continue
        try:
            # Format: "Mon\n4/6"
            parts = header.split("\n")
            month, day = parts[1].split("/")
            col_dates[col] = date(2026, int(month), int(day))
        except Exception:
            continue

    categories = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v:
            categories.append((row, str(v).strip()))

    # Build records: {date: {category: seconds}}
    records = defaultdict(lambda: defaultdict(float))
    for row, cat in categories:
        for col, d in col_dates.items():
            v = ws.cell(row, col).value
            if isinstance(v, timedelta):
                secs = v.total_seconds()
            elif isinstance(v, (int, float)) and v > 0:
                secs = v * 86400
            else:
                secs = 0.0
            if secs > 0:
                records[d][cat] = secs

    return records, [c for _, c in categories]


def fmt(seconds):
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h}h {m:02}m {sec:02}s"


def fmt_h(seconds):
    return f"{seconds/3600:.2f}h"


def section(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


# ── Main analysis ─────────────────────────────────────────────

def run(path=EXCEL_FILE):
    records, categories = load_data(path)

    if not records:
        print("No data found.")
        return

    all_dates = sorted(records.keys())
    start, end = all_dates[0], all_dates[-1]
    active_days = len(all_dates)

    # Build a flat DataFrame: date, category, seconds
    rows = []
    for d, cats in records.items():
        for cat, secs in cats.items():
            rows.append({"date": d, "category": cat, "seconds": secs, "hours": secs / 3600})
    df = pd.DataFrame(rows)
    df["dow"] = df["date"].apply(lambda d: d.strftime("%A"))
    df["week"] = df["date"].apply(lambda d: d.isocalendar()[1])

    # Daily totals
    daily = df.groupby("date")["seconds"].sum().reset_index()
    daily.columns = ["date", "total_seconds"]
    daily["dow"] = daily["date"].apply(lambda d: d.strftime("%A"))

    # ── 1. Overview ───────────────────────────────────────────
    section("OVERVIEW")
    total_secs = df["seconds"].sum()
    print(f"  Tracking period : {start} → {end}")
    print(f"  Active days     : {active_days}")
    print(f"  Total tracked   : {fmt(total_secs)}")
    print(f"  Daily avg       : {fmt(total_secs / active_days)} (on active days)")

    # ── 2. Totals per category ────────────────────────────────
    section("TOTAL TIME PER CATEGORY")
    cat_totals = df.groupby("category")["seconds"].sum().sort_values(ascending=False)
    for cat, secs in cat_totals.items():
        pct = secs / total_secs * 100
        bar = "█" * int(pct / 2)
        print(f"  {cat:<20} {fmt(secs):>18}  ({pct:5.1f}%)  {bar}")

    # ── 3. Daily average per category ────────────────────────
    section("DAILY AVERAGE PER CATEGORY  (on days you tracked it)")
    cat_days = df.groupby("category")["date"].nunique()
    cat_avg = cat_totals / cat_days
    for cat in cat_totals.index:
        print(f"  {cat:<20} avg {fmt_h(cat_avg[cat]):>8} / day  ({cat_days[cat]} days tracked)")

    # ── 4. Day-of-week productivity ───────────────────────────
    section("PRODUCTIVITY BY DAY OF WEEK")
    DOW_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    dow_avg = daily.groupby("dow")["total_seconds"].mean()
    dow_count = daily.groupby("dow")["total_seconds"].count()
    print(f"  {'Day':<12} {'Avg Hours':>10}  {'Sessions':>9}  Chart")
    for dow in DOW_ORDER:
        if dow not in dow_avg:
            continue
        avg_h = dow_avg[dow] / 3600
        bar = "█" * int(avg_h * 2)
        print(f"  {dow:<12} {avg_h:>9.2f}h  {int(dow_count[dow]):>9}  {bar}")

    # ── 5. Category breakdown by day of week ─────────────────
    section("CATEGORY BREAKDOWN BY DAY OF WEEK  (avg hours)")
    cat_dow = df.groupby(["category", "dow"])["hours"].mean().unstack(fill_value=0)
    # Reorder columns
    available_days = [d for d in DOW_ORDER if d in cat_dow.columns]
    cat_dow = cat_dow[available_days]
    header = f"  {'Category':<22}" + "".join(f"{d[:3]:>8}" for d in available_days)
    print(header)
    print("  " + "-" * (22 + 8 * len(available_days)))
    for cat in cat_totals.index:
        row_str = f"  {cat:<22}"
        for dow in available_days:
            val = cat_dow.loc[cat, dow] if cat in cat_dow.index and dow in cat_dow.columns else 0
            row_str += f"{val:>7.2f}h" if val > 0 else f"{'—':>8}"
        print(row_str)

    # ── 6. Streaks ────────────────────────────────────────────
    section("STREAKS & CONSISTENCY")
    # Overall: any activity
    date_range = pd.date_range(start, end).date
    active_set = set(daily["date"])

    def streak_stats(active):
        cur_streak = max_streak = 0
        last_d = None
        for d in sorted(date_range):
            if d in active:
                if last_d and (d - last_d).days == 1:
                    cur_streak += 1
                else:
                    cur_streak = 1
                max_streak = max(max_streak, cur_streak)
                last_d = d
            else:
                last_d = None
                cur_streak = 0
        return max_streak, cur_streak

    max_s, cur_s = streak_stats(active_set)
    print(f"\n  Overall (any category)")
    print(f"    Longest streak  : {max_s} consecutive days")
    print(f"    Current streak  : {cur_s} days")
    print(f"    Active rate     : {active_days / len(date_range) * 100:.1f}% of days in range")

    print(f"\n  Per-category streaks:")
    for cat in cat_totals.index:
        cat_active = set(df[df["category"] == cat]["date"])
        max_s, cur_s = streak_stats(cat_active)
        active_d = len(cat_active)
        print(f"    {cat:<22} longest={max_s}d  current={cur_s}d  active={active_d} days")

    # ── 7. Category tradeoffs ─────────────────────────────────
    section("CATEGORY TRADEOFFS  (correlation on shared days)")
    print("  When you spend more time on X, how does Y change?")
    print("  Positive = they go together | Negative = they compete\n")

    # Pivot: date × category hours
    pivot = df.pivot_table(index="date", columns="category", values="hours", fill_value=0)
    corr = pivot.corr()

    # Print upper triangle only
    cats_list = list(corr.columns)
    print(f"  {'':28}", end="")
    for c in cats_list:
        print(f"{c[:8]:>10}", end="")
    print()
    print("  " + "-" * (28 + 10 * len(cats_list)))
    for i, c1 in enumerate(cats_list):
        print(f"  {c1:<28}", end="")
        for j, c2 in enumerate(cats_list):
            if j < i:
                print(f"{'':>10}", end="")
            elif j == i:
                print(f"{'1.00':>10}", end="")
            else:
                val = corr.loc[c1, c2]
                print(f"{val:>10.2f}", end="")
        print()

    # Highlight strongest negative correlations (real tradeoffs)
    print(f"\n  Strongest tradeoffs (most negative correlations):")
    pairs = []
    for i, c1 in enumerate(cats_list):
        for j, c2 in enumerate(cats_list):
            if j > i:
                pairs.append((corr.loc[c1, c2], c1, c2))
    pairs.sort()
    for val, c1, c2 in pairs[:5]:
        direction = "compete ↕" if val < -0.1 else "neutral  " if abs(val) < 0.2 else "go together ↑"
        print(f"    {c1} ↔ {c2:<20} r={val:+.2f}  {direction}")

    # ── 8. Weekly totals ──────────────────────────────────────
    section("WEEKLY TOTALS")
    daily["week"] = daily["date"].apply(lambda d: d.isocalendar()[1])
    weekly = daily.groupby("week")["total_seconds"].sum()
    for week_num, secs in weekly.items():
        bar = "█" * int(secs / 3600)
        print(f"  Week {week_num:>2}: {fmt_h(secs):>8}  {bar}")

    print(f"\n{'='*60}\n")


if __name__ == "__main__":
    run()